"""Pure parser and deterministic rules for T-Bank XLSX statements."""

from __future__ import annotations

import hashlib
from dataclasses import dataclass
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from io import BytesIO
from typing import Literal

from openpyxl import load_workbook

MONEY_QUANTUM = Decimal("0.01")
REQUIRED_HEADERS = (
    "тип операции",
    "дата проведения",
    "номер документа",
    "сумма в валюте счета",
    "описание операции",
    "назначение платежа",
    "наименование контрагента",
    "инн контрагента",
)
BASE_HEADERS = REQUIRED_HEADERS[:6]
COUNTERPARTY_NAME = "наименование контрагента"
COUNTERPARTY_INN = "инн контрагента"
PAYER_NAME = "наименование плательщика"
PAYER_INN = "инн плательщика"
RECIPIENT_NAME = "наименование получателя"
RECIPIENT_INN = "инн получателя"
KNOWN_HEADERS = frozenset(
    (*REQUIRED_HEADERS, PAYER_NAME, PAYER_INN, RECIPIENT_NAME, RECIPIENT_INN)
)
HEADER_SCAN_LIMIT = 40


class BankImportError(ValueError):
    """A safe validation error that never contains statement row values."""

    def __init__(
        self,
        message: str,
        *,
        found_columns: list[str] | None = None,
        missing_columns: list[str] | None = None,
        searched_row: int | None = None,
    ) -> None:
        super().__init__(message)
        self.found_columns = found_columns or []
        self.missing_columns = missing_columns or []
        self.searched_row = searched_row

    def as_detail(self) -> dict[str, str | int | list[str] | None]:
        return {
            "message": str(self),
            "found_columns": self.found_columns,
            "missing_columns": self.missing_columns,
            "searched_row": self.searched_row,
        }


@dataclass(frozen=True)
class BankRow:
    operation_type: str
    operation_date: date
    doc_number: str
    amount: Decimal
    description: str
    payment_purpose: str
    counterparty_name: str
    counterparty_inn: str


@dataclass(frozen=True)
class ClassificationResult:
    kind: Literal["income", "expense"]
    category: str | None
    channel: str | None
    needs_review: bool
    is_transfer: bool


def parse_amount(value: object) -> Decimal:
    """Parse a spreadsheet monetary value without binary-float arithmetic."""
    if isinstance(value, bool) or value is None:
        raise BankImportError("invalid amount")

    if isinstance(value, (int, float, Decimal)):
        raw = str(value)
    elif isinstance(value, str):
        raw = value.strip().replace("\u00a0", "").replace("\u202f", "")
        raw = raw.replace(" ", "")
        if not raw:
            raise BankImportError("invalid amount")
        comma_index = raw.rfind(",")
        dot_index = raw.rfind(".")
        if comma_index >= 0 and dot_index >= 0:
            decimal_separator = "," if comma_index > dot_index else "."
            thousands_separator = "." if decimal_separator == "," else ","
            raw = raw.replace(thousands_separator, "")
            if decimal_separator == ",":
                raw = raw.replace(",", ".")
        elif comma_index >= 0:
            raw = raw.replace(",", ".")
    else:
        raise BankImportError("invalid amount")

    try:
        return Decimal(raw).quantize(MONEY_QUANTUM, rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError) as error:
        raise BankImportError("invalid amount") from error


def parse_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        raw = value.strip()
        for pattern in ("%Y-%m-%d", "%d.%m.%Y"):
            try:
                return datetime.strptime(raw, pattern).date()
            except ValueError:
                continue
    raise BankImportError("invalid operation date")


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_header(value: object) -> str:
    if not isinstance(value, str):
        return ""
    return " ".join(value.strip().lower().replace("ё", "е").split())


def _header_map(values: tuple[object, ...]) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, value in enumerate(values):
        normalized = _normalize_header(value)
        if normalized:
            result[normalized] = index
    return result


def _missing_headers(indexes: dict[str, int]) -> list[str]:
    missing = [header for header in BASE_HEADERS if header not in indexes]
    if COUNTERPARTY_NAME not in indexes and not all(
        header in indexes for header in (PAYER_NAME, RECIPIENT_NAME)
    ):
        missing.append(COUNTERPARTY_NAME)
    if COUNTERPARTY_INN not in indexes and not all(
        header in indexes for header in (PAYER_INN, RECIPIENT_INN)
    ):
        missing.append(COUNTERPARTY_INN)
    return missing


def _directional_value(
    values: tuple[object, ...],
    indexes: dict[str, int],
    operation_type: str,
    direct_header: str,
    payer_header: str,
    recipient_header: str,
) -> object:
    header = direct_header
    if header not in indexes:
        direction = operation_type.strip().casefold().replace("ё", "е")
        header = payer_header if direction == "кредит" else recipient_header
    index = indexes[header]
    return values[index] if index < len(values) else None


def parse_tbank_xlsx(content: bytes) -> list[BankRow]:
    """Read the first sheet containing the complete T-Bank header set."""
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as error:
        raise BankImportError("invalid xlsx file") from error

    best_found: list[str] = []
    best_missing = list(REQUIRED_HEADERS)
    best_row: int | None = None
    try:
        for sheet in workbook.worksheets:
            # Some T-Bank exports declare the worksheet range as A1 even though
            # rows continue below it. Read-only mode trusts that metadata unless
            # dimensions are reset, which would otherwise hide the statement.
            sheet.reset_dimensions()
            header_indexes: dict[str, int] | None = None
            header_row_number = 0
            for row_number, values in enumerate(
                sheet.iter_rows(values_only=True), start=1
            ):
                if row_number > HEADER_SCAN_LIMIT:
                    break
                normalized = _header_map(values)
                found = [header for header in normalized if header in KNOWN_HEADERS]
                missing = _missing_headers(normalized)
                if len(found) > len(best_found):
                    best_found = found
                    best_missing = missing
                    best_row = row_number
                if not missing:
                    header_indexes = normalized
                    header_row_number = row_number
                    break

            if header_indexes is None:
                continue

            result: list[BankRow] = []
            for row_number, values in enumerate(
                sheet.iter_rows(values_only=True), start=1
            ):
                if row_number <= header_row_number:
                    continue
                if all(value in (None, "") for value in values):
                    continue

                def value_for(
                    header: str,
                    indexes: dict[str, int] = header_indexes,
                    row_values: tuple[object, ...] = values,
                ) -> object:
                    index = indexes[header]
                    return row_values[index] if index < len(row_values) else None

                try:
                    operation_type = _cell_text(value_for("тип операции"))
                    result.append(
                        BankRow(
                            operation_type=operation_type,
                            operation_date=parse_date(value_for("дата проведения")),
                            doc_number=_cell_text(value_for("номер документа")),
                            amount=parse_amount(value_for("сумма в валюте счета")),
                            description=_cell_text(value_for("описание операции")),
                            payment_purpose=_cell_text(value_for("назначение платежа")),
                            counterparty_name=_cell_text(
                                _directional_value(
                                    values,
                                    header_indexes,
                                    operation_type,
                                    COUNTERPARTY_NAME,
                                    PAYER_NAME,
                                    RECIPIENT_NAME,
                                )
                            ),
                            counterparty_inn=_cell_text(
                                _directional_value(
                                    values,
                                    header_indexes,
                                    operation_type,
                                    COUNTERPARTY_INN,
                                    PAYER_INN,
                                    RECIPIENT_INN,
                                )
                            ),
                        )
                    )
                except BankImportError as error:
                    raise BankImportError(
                        f"invalid statement row {row_number}"
                    ) from error
            return result
    finally:
        workbook.close()

    raise BankImportError(
        "required columns not found",
        found_columns=best_found,
        missing_columns=best_missing,
        searched_row=best_row,
    )


def source_hash(row: BankRow) -> str:
    canonical = (
        f"{row.operation_date.isoformat()}|{row.amount.quantize(MONEY_QUANTUM):.2f}|"
        f"{row.doc_number.strip()}|{row.counterparty_inn.strip()}"
    )
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def mask_inn(value: str) -> str:
    normalized = value.strip()
    if not normalized:
        return ""
    if len(normalized) <= 4:
        return "*" * len(normalized)
    return "*" * (len(normalized) - 4) + normalized[-4:]


def _contains(text: str, *needles: str) -> bool:
    folded = text.casefold()
    return any(needle.casefold() in folded for needle in needles)


def classify_transaction(row: BankRow) -> ClassificationResult:
    operation_type = row.operation_type.strip().casefold()
    purpose = row.payment_purpose
    description = row.description
    counterparty = row.counterparty_name

    if operation_type == "кредит":
        if _contains(purpose, "перевод собственных средств", "возврат"):
            return ClassificationResult("income", None, None, False, True)
        if _contains(purpose, "акарицид", "клещ"):
            return ClassificationResult(
                "income", "Обработка от клещей", "Прочее", False, False
            )
        if _contains(purpose, "химчистк"):
            return ClassificationResult("income", "Химчистка", "Прочее", False, False)
        if _contains(purpose, "дератизац"):
            return ClassificationResult("income", "Дератизация", "Прочее", False, False)
        if _contains(purpose, "дезинфекц"):
            return ClassificationResult("income", "Дезинфекция", "Прочее", False, False)
        if _contains(purpose, "дезинсекц", "таракан", "клоп"):
            return ClassificationResult("income", "Дезинсекция", "Прочее", False, False)
        if row.counterparty_inn.strip():
            return ClassificationResult(
                "income", "Юридические клиенты", "Прочее", True, False
            )
        return ClassificationResult("income", "Другие работы", None, True, False)

    if operation_type == "дебет":
        if _contains(purpose, "перевод собственных средств"):
            return ClassificationResult("expense", None, None, False, True)
        if _contains(counterparty, "медилис") or _contains(description, "дез средства"):
            return ClassificationResult(
                "expense", "Материалы и химия", None, False, False
            )
        if _contains(f"{purpose} {description}", "енс", "фнс", "взносы"):
            return ClassificationResult(
                "expense", "Налоги и взносы", None, False, False
            )
        if _contains(counterparty, "тбанк") or _contains(
            f"{purpose} {description}", "комиссия", "sms", "обслуживание"
        ):
            return ClassificationResult(
                "expense", "Банковские комиссии", None, False, False
            )
        return ClassificationResult("expense", None, None, True, False)

    raise BankImportError("unsupported operation type")


def transaction_comment(row: BankRow) -> str:
    return row.payment_purpose.strip() or row.description.strip()
