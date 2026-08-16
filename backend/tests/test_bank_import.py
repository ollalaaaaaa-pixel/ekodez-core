import hashlib
import re
import unittest
import zipfile
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, cast

from openpyxl import Workbook

from app.bank_import import (
    BankImportError,
    BankRow,
    classify_transaction,
    mask_inn,
    parse_amount,
    parse_date,
    parse_tbank_xlsx,
    source_hash,
    transaction_comment,
)

HEADERS = (
    "Тип операции",
    "Дата проведения",
    "Номер документа",
    "Сумма в валюте счёта",
    "Описание операции",
    "Назначение платежа",
    "Наименование контрагента",
    "ИНН контрагента",
)

REAL_LAYOUT_HEADERS = (
    " ТИП   ОПЕРАЦИИ ",
    "Дата проведения",
    "Номер документа",
    "Сумма в валюте СЧЁТА",
    "Описание операции",
    "Назначение платежа",
    "Наименование плательщика",
    "ИНН плательщика",
    "Наименование получателя",
    "ИНН получателя",
    "Счёт плательщика",
    "Счёт получателя",
    "БИК плательщика",
    "БИК получателя",
    "Банк плательщика",
    "Банк получателя",
    "КПП плательщика",
    "КПП получателя",
    "Валюта",
    "Статус",
    "Очередность",
    "Код операции",
    "УИН",
    "НДС",
    "КБК",
    "ОКТМО",
    "Период",
    "Служебная колонка",
)


def make_row(**overrides: object) -> BankRow:
    values: dict[str, object] = {
        "operation_type": "Кредит",
        "operation_date": date(2026, 8, 15),
        "doc_number": " 42 ",
        "amount": Decimal("1500.50"),
        "description": "Описание",
        "payment_purpose": "Оплата услуг",
        "counterparty_name": "ООО Синтетика",
        "counterparty_inn": " 2901000000 ",
    }
    values.update(overrides)
    return BankRow(**values)  # type: ignore[arg-type]


def workbook_bytes(*rows: tuple[object, ...], include_preamble: bool = True) -> bytes:
    workbook = Workbook()
    sheet = workbook.worksheets[0]
    if include_preamble:
        sheet.append(("Выписка",))
        sheet.append(("Служебная строка",))
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(cast(list[Any], list(row)))
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def real_layout_bytes(
    *rows: tuple[object, ...], headers: tuple[str, ...] = REAL_LAYOUT_HEADERS
) -> bytes:
    workbook = Workbook()
    sheet = workbook.worksheets[0]
    for label in (
        "Выписка по счёту №",
        "Клиент",
        "ИНН",
        "КПП",
        "Входящий остаток",
        "Исходящий остаток",
        "Обороты по дебету",
        "Обороты по кредиту",
    ):
        sheet.append((label,))
    sheet.append(())
    sheet.append(headers)
    for row in rows:
        sheet.append(cast(list[Any], list(row)))
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def real_layout_row(
    operation_type: str,
    payer_name: str,
    payer_inn: str,
    recipient_name: str,
    recipient_inn: str,
) -> tuple[object, ...]:
    return (
        operation_type,
        "15.08.2025",
        "DOC-1",
        "27,544.00",
        "Описание",
        "Назначение",
        payer_name,
        payer_inn,
        recipient_name,
        recipient_inn,
        *("" for _ in range(18)),
    )


def force_incorrect_a1_dimension(content: bytes) -> bytes:
    source = BytesIO(content)
    output = BytesIO()
    with (
        zipfile.ZipFile(source) as source_zip,
        zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as output_zip,
    ):
        for item in source_zip.infolist():
            data = source_zip.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                data, replacements = re.subn(
                    rb'<dimension ref="[^"]+"\s*/>', b'<dimension ref="A1"/>', data
                )
                if replacements != 1:
                    raise AssertionError("worksheet dimension fixture was not replaced")
            output_zip.writestr(item, data)
    return output.getvalue()


class AmountAndDateParsingTest(unittest.TestCase):
    def test_string_amounts_support_russian_and_grouped_dot_formats(self):
        self.assertEqual(parse_amount("1 500 000,50"), Decimal("1500000.50"))
        self.assertEqual(parse_amount("27,544.00"), Decimal("27544.00"))
        self.assertEqual(parse_amount("1\u00a0234\u202f567,895"), Decimal("1234567.90"))

    def test_numeric_cells_use_decimal_string_conversion(self):
        self.assertEqual(parse_amount(1500.5), Decimal("1500.50"))
        self.assertEqual(parse_amount(Decimal("0.005")), Decimal("0.01"))

    def test_invalid_amount_is_rejected_without_echoing_value(self):
        with self.assertRaisesRegex(BankImportError, "invalid amount") as error:
            parse_amount("секретное значение")
        self.assertNotIn("секретное", str(error.exception))

    def test_dates_accept_excel_datetime_date_and_known_strings(self):
        self.assertEqual(parse_date(datetime(2026, 8, 15, 12, 30)), date(2026, 8, 15))
        self.assertEqual(parse_date(date(2026, 8, 16)), date(2026, 8, 16))
        self.assertEqual(parse_date("17.08.2026"), date(2026, 8, 17))
        self.assertEqual(parse_date("2026-08-18"), date(2026, 8, 18))


class WorkbookParsingTest(unittest.TestCase):
    def test_real_statement_with_incorrect_a1_dimension_is_still_read(self):
        content = force_incorrect_a1_dimension(
            real_layout_bytes(
                real_layout_row(
                    "Кредит",
                    "Плательщик CREDIT",
                    "1111111111",
                    "Получатель CREDIT",
                    "2222222222",
                )
            )
        )

        rows = parse_tbank_xlsx(content)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].amount, Decimal("27544.00"))

    def test_real_layout_normalizes_headers_and_uses_directional_fallback(self):
        content = real_layout_bytes(
            real_layout_row(
                "Кредит",
                "Плательщик CREDIT",
                "1111111111",
                "Получатель CREDIT",
                "2222222222",
            ),
            real_layout_row(
                "Дебет",
                "Плательщик DEBIT",
                "3333333333",
                "Получатель DEBIT",
                "4444444444",
            ),
        )

        rows = parse_tbank_xlsx(content)

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0].counterparty_name, "Плательщик CREDIT")
        self.assertEqual(rows[0].counterparty_inn, "1111111111")
        self.assertEqual(rows[1].counterparty_name, "Получатель DEBIT")
        self.assertEqual(rows[1].counterparty_inn, "4444444444")
        self.assertEqual(rows[0].amount, Decimal("27544.00"))
        self.assertEqual(rows[0].operation_date, date(2025, 8, 15))

    def test_exact_counterparty_columns_win_over_similar_columns(self):
        headers = list(REAL_LAYOUT_HEADERS)
        headers[10] = "Наименование контрагента"
        headers[11] = "ИНН контрагента"
        row = list(
            real_layout_row(
                "Кредит",
                "Плательщик WRONG",
                "1111111111",
                "Получатель WRONG",
                "2222222222",
            )
        )
        row[10] = "Контрагент EXACT"
        row[11] = "5555555555"

        parsed = parse_tbank_xlsx(real_layout_bytes(tuple(row), headers=tuple(headers)))

        self.assertEqual(parsed[0].counterparty_name, "Контрагент EXACT")
        self.assertEqual(parsed[0].counterparty_inn, "5555555555")

    def test_missing_headers_report_best_row_found_and_missing_columns(self):
        headers = list(REAL_LAYOUT_HEADERS)
        headers[7] = "Неизвестная колонка 1"
        headers[9] = "Неизвестная колонка 2"

        with self.assertRaises(BankImportError) as context:
            parse_tbank_xlsx(real_layout_bytes(headers=tuple(headers)))

        self.assertEqual(context.exception.searched_row, 10)
        self.assertIn("тип операции", context.exception.found_columns)
        self.assertIn("инн контрагента", context.exception.missing_columns)
        self.assertNotIn("инн плательщика", context.exception.found_columns)

    def test_parser_finds_headers_after_preamble_and_reads_rows(self):
        content = workbook_bytes(
            (
                "Кредит",
                datetime(2026, 8, 15, 10, 0),
                101,
                1500.5,
                "Описание 1",
                "Назначение 1",
                "Контрагент 1",
                2901000000,
            ),
            (
                "Дебет",
                "16.08.2026",
                " 102 ",
                "27,544.00",
                "Описание 2",
                "",
                "Контрагент 2",
                "",
            ),
        )

        rows = parse_tbank_xlsx(content)

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0].operation_type, "Кредит")
        self.assertEqual(rows[0].operation_date, date(2026, 8, 15))
        self.assertEqual(rows[0].doc_number, "101")
        self.assertEqual(rows[0].amount, Decimal("1500.50"))
        self.assertEqual(rows[0].counterparty_inn, "2901000000")
        self.assertEqual(rows[1].operation_date, date(2026, 8, 16))
        self.assertEqual(rows[1].amount, Decimal("27544.00"))

    def test_missing_required_column_is_rejected(self):
        workbook = Workbook()
        workbook.worksheets[0].append(HEADERS[:-1])
        output = BytesIO()
        workbook.save(output)

        with self.assertRaisesRegex(BankImportError, "required columns not found"):
            parse_tbank_xlsx(output.getvalue())

    def test_corrupted_workbook_is_rejected_without_library_details(self):
        with self.assertRaisesRegex(BankImportError, "invalid xlsx file"):
            parse_tbank_xlsx(b"not-a-workbook")


class HashAndMaskTest(unittest.TestCase):
    def test_hash_uses_canonical_trimmed_fields_and_two_decimal_amount(self):
        row = make_row()
        expected = hashlib.sha256(b"2026-08-15|1500.50|42|2901000000").hexdigest()

        self.assertEqual(source_hash(row), expected)
        self.assertEqual(len(source_hash(row)), 64)

    def test_inn_is_masked_for_ui_and_logs(self):
        self.assertEqual(mask_inn("2901000000"), "******0000")
        self.assertEqual(mask_inn("123"), "***")
        self.assertEqual(mask_inn(""), "")


class ClassificationTest(unittest.TestCase):
    def test_credit_rules_follow_approved_precedence(self):
        cases = (
            ("Перевод собственных средств", "", "", None, None, False, True),
            ("Возврат по ПП", "", "", None, None, False, True),
            (
                "акарицид и дезинсекция",
                "",
                "",
                "Обработка от клещей",
                "Прочее",
                False,
                False,
            ),
            ("оплата химчистки", "", "", "Химчистка", "Прочее", False, False),
            ("дератизация", "", "", "Дератизация", "Прочее", False, False),
            ("дезинфекция", "", "", "Дезинфекция", "Прочее", False, False),
            ("вывод тараканов", "", "", "Дезинсекция", "Прочее", False, False),
            (
                "обычная услуга",
                "",
                "2901000000",
                "Юридические клиенты",
                "Прочее",
                True,
                False,
            ),
            ("обычная услуга", "", "", "Другие работы", None, True, False),
        )
        for purpose, description, inn, category, channel, review, transfer in cases:
            with self.subTest(purpose=purpose, inn=bool(inn)):
                result = classify_transaction(
                    make_row(
                        payment_purpose=purpose,
                        description=description,
                        counterparty_inn=inn,
                    )
                )
                self.assertEqual(result.kind, "income")
                self.assertEqual(result.category, category)
                self.assertEqual(result.channel, channel)
                self.assertEqual(result.needs_review, review)
                self.assertEqual(result.is_transfer, transfer)

    def test_debit_rules_follow_approved_precedence(self):
        cases = (
            ("Перевод собственных средств", "", "", None, False, True),
            ("Оплата", "", "МЕДИЛИС", "Материалы и химия", False, False),
            ("Оплата", "Дез средства", "Поставщик", "Материалы и химия", False, False),
            ("Оплата ЕНС", "", "Поставщик", "Налоги и взносы", False, False),
            (
                "Оплата",
                "Фиксированные взносы",
                "Поставщик",
                "Налоги и взносы",
                False,
                False,
            ),
            (
                "Оплата",
                "Комиссия за обслуживание",
                "ТБанк",
                "Банковские комиссии",
                False,
                False,
            ),
            ("Оплата", "Неизвестно", "Поставщик", None, True, False),
        )
        for purpose, description, counterparty, category, review, transfer in cases:
            with self.subTest(purpose=purpose, description=description):
                result = classify_transaction(
                    make_row(
                        operation_type="Дебет",
                        payment_purpose=purpose,
                        description=description,
                        counterparty_name=counterparty,
                        counterparty_inn="",
                    )
                )
                self.assertEqual(result.kind, "expense")
                self.assertEqual(result.category, category)
                self.assertIsNone(result.channel)
                self.assertEqual(result.needs_review, review)
                self.assertEqual(result.is_transfer, transfer)

    def test_comment_uses_purpose_then_description(self):
        self.assertEqual(transaction_comment(make_row()), "Оплата услуг")
        self.assertEqual(
            transaction_comment(make_row(payment_purpose="  ", description=" Резерв ")),
            "Резерв",
        )

    def test_unknown_operation_type_is_rejected(self):
        with self.assertRaisesRegex(BankImportError, "unsupported operation type"):
            classify_transaction(make_row(operation_type="Другое"))


if __name__ == "__main__":
    unittest.main()
